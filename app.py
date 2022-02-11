'''
import openpyxl as xl
from openpyxl.chart import BarChart, Reference
wb = xl.load_workbook('excel_work.xlsx')
sheet = wb['Raw Data']
cell = sheet.cell(1, 1)


for row in range(1, 664):
    cell = sheet.cell(row, 16)
    corrected_price = cell.value * 10
    corrected_values_cell = sheet.cell(row, 17)
    corrected_values_cell.value = corrected_price

values = Reference(sheet, min_row=2, max_row=3, min_col=3, max_col=3)
chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart, 'x2')

wb.save('excel_work1.xlsx')
'''


F_name = input('What is your first name? ')
F_name.upper = input('What is your first name? ')
L_name = input('what is your last name? ')
Birth_year = input('what is your year of birth? ')
age = (2022 - int(Birth_year))
print(age + 3)
if age + 3 <=30:
    print('You are a young man from Ghana')
else:
    print('You should be in University by now and that is great achievement.')
